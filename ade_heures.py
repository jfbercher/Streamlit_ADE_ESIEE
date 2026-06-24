#!/usr/bin/env python3
"""
ade_heures.py - Parse ADE-exported .ics files and generate an Excel breakdown of teaching hours.

Usage: python ade_heures.py input.ics [output.xlsx]
"""

import re
import sys
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from collections import defaultdict
import pandas as pd

_PARIS_TZ = ZoneInfo("Europe/Paris")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# ICS Parsing
# ---------------------------------------------------------------------------

def unfold_lines(raw_text):
    """Unfold RFC 5545 folded lines (continuation lines start with a space or tab)."""
    lines = raw_text.splitlines()
    unfolded = []
    for line in lines:
        if line and line[0] in (' ', '\t'):
            if unfolded:
                unfolded[-1] += line[1:]
            else:
                unfolded.append(line.lstrip())
        else:
            unfolded.append(line)
    return unfolded


def parse_ics(filepath):
    """Parse an ICS file and return a list of VEVENT property dicts."""
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        raw = f.read()

    lines = unfold_lines(raw)
    events = []
    current = None

    for line in lines:
        if line == 'BEGIN:VEVENT':
            current = {}
        elif line == 'END:VEVENT':
            if current is not None:
                events.append(current)
            current = None
        elif current is not None and ':' in line:
            key, _, value = line.partition(':')
            # Strip property parameters (e.g. DTSTART;TZID=Europe/Paris)
            key_base = key.split(';')[0].strip()
            current[key_base] = value.strip()

    return events


def parse_dt(dt_str):
    """Parse a DTSTART/DTEND value; convert UTC (Z suffix) to Paris time (UTC+1)."""
    dt_str = dt_str.strip()
    is_utc = dt_str.endswith('Z')
    if is_utc:
        dt_str = dt_str[:-1]
    try:
        dt = datetime.strptime(dt_str, '%Y%m%dT%H%M%S')
    except ValueError:
        try:
            dt = datetime.strptime(dt_str, '%Y%m%d')
        except ValueError:
            return None
    if is_utc:
        # Convertir UTC → Europe/Paris (gère CET +1h en hiver et CEST +2h en été automatiquement)
        dt = datetime.fromtimestamp(dt.replace(tzinfo=timezone.utc).timestamp(), tz=_PARIS_TZ).replace(tzinfo=None)
    return dt


def decode_description(desc_raw):
    """
    Decode ADE's DESCRIPTION field.
    ADE encodes newlines as literal \\n sequences inside the property value.
    Returns a list of non-empty stripped lines.
    """
    text = desc_raw.replace('\\n', '\n').replace('\\,', ',')
    lines = [l.strip() for l in text.split('\n')]
    return [l for l in lines if l]


# ---------------------------------------------------------------------------
# Filière / Promo Extraction
# ---------------------------------------------------------------------------

_CODE_RE  = re.compile(r'^(?:[A-Z0-9][A-Z0-9\-]*|E\d[A-Za-z0-9\-]*)$')  # majuscules, ou E+chiffre avec minuscules possibles (ex: E3-1d-S1)
_PROMO_RE = re.compile(r'^E\d')                                            # starts with E + digit → promo level


def extract_codes(desc_lines):
    """
    Extract filière and promo codes from the beginning of desc_lines.

    ADE places group/filière codes at the top of the DESCRIPTION, before the
    course name or teacher name (which contain spaces).  Codes are compact
    (no spaces): all-caps filières (BIO, AIC…) or E+digit promos that may
    contain lowercase letters (E3-1d-S1, E3-1s-S1…).  We collect them until
    the first line that does not match _CODE_RE, then split into:
      - promos  : codes matching ^E\\d  (e.g. E2-5-EST, E3-1d-S1)
      - filieres: everything else        (e.g. BIO, AIC, DSIA, GI)

    Returns (promos: list[str], filieres: list[str]) — both deduplicated,
    order preserved.
    """
    seen, all_codes = set(), []
    for line in desc_lines:
        if _CODE_RE.match(line):
            if line not in seen:
                seen.add(line)
                all_codes.append(line)
        else:
            break  # first line with lowercase → end of code block
    promos   = [c for c in all_codes if _PROMO_RE.match(c)]
    filieres = [c for c in all_codes if not _PROMO_RE.match(c)]
    return promos, filieres


# ---------------------------------------------------------------------------
# Modality Detection
# ---------------------------------------------------------------------------

# Priority order matters: more specific patterns first.
MODALITY_PATTERNS = [
    ('CM/TD',   re.compile(r'\bCM[/-]TD\b',        re.IGNORECASE)),
    ('Oraux',   re.compile(r'\bOraux\b',         re.IGNORECASE)),
    ('TDR',     re.compile(r'\bTDR\b',           re.IGNORECASE)),
    ('TP Seul', re.compile(r'\bTP\s+[Ss]eul\b',  re.IGNORECASE)),
    ('CM',      re.compile(r'\bCM\d*\b')),
    ('TD',      re.compile(r'\bTD\d*\b')),
    ('TP',      re.compile(r'\bTP\d*\b')),
    ('Soutenance',   re.compile(r'\bSoutenance\b',        re.IGNORECASE)),
    ('CM-plus',   re.compile(r'\bCM\-plus\b',        re.IGNORECASE)),
    
]


def detect_modality(desc_lines):
    """
    Detect modality from the LAST non-empty line of the description only.
    Using only the last line avoids false positives from course names,
    teacher names, or room codes that may contain 'TP', 'CM', 'TD', etc.
    """
    if not desc_lines:
        return 'Autre'
    last_line = desc_lines[-1]
    for modality, pattern in MODALITY_PATTERNS:
        if pattern.search(last_line):
            return modality
    return 'Autre'


# ---------------------------------------------------------------------------
# Event Processing
# ---------------------------------------------------------------------------

def process_events(raw_events):
    """Convert raw ICS property dicts into structured event records."""
    records = []
    for ev in raw_events:
        summary = ev.get('SUMMARY', '').replace('\\,', ',').replace('\\n', ' ').strip()
        location = ev.get('LOCATION', '').replace('\\,', ',').strip()
        desc_raw = ev.get('DESCRIPTION', '')
        desc_lines = decode_description(desc_raw)
        description_clean = ' | '.join(desc_lines)

        dtstart = parse_dt(ev.get('DTSTART', ''))
        dtend = parse_dt(ev.get('DTEND', ''))

        if dtstart is None or dtend is None:
            continue

        duration_h = (dtend - dtstart).total_seconds() / 3600.0
        if duration_h <= 0:
            continue

        promos, filieres = extract_codes(desc_lines)
        modality = detect_modality(desc_lines)

        # Fallback : si la description ne donne aucune modalité, chercher dans le nom du cours
        # On suffixe _Trou_ADE pour signaler que la détection est incertaine (à vérifier)
        if modality == 'Autre' and summary:
            for m, pattern in MODALITY_PATTERNS:
                if pattern.search(summary):
                    modality = m + '_Trou_ADE'
                    print(modality)
                    break

        # CM+ si : promo E1/E2, OU 5 groupes ou plus (toutes années confondues)
        base_mod = modality.replace("_Trou_ADE", "")
        # ^E[12](-[A-Za-z]+)*$ : promotion entière (E2, E2-BIO…) mais pas un groupe (E2-5-EST, E3-1d-S1)
        is_cm_plus = base_mod == 'CM' and (
            any(re.match(r'^E[12](-[A-Za-z]+)*$', p) for p in promos) or len(promos) >= 5
        )
        if is_cm_plus:
            modality = 'CM-plus_Trou_ADE' if '_Trou_ADE' in modality else 'CM-plus'
        records.append({
            'nom':         summary,
            'dtstart':     dtstart,
            'dtend':       dtend,
            'duration_h':  duration_h,
            'location':    location,
            'modality':    modality,
            'promo':       ' / '.join(promos)   or '—',
            'filiere':     ' / '.join(filieres) or '—',
            'description': description_clean,
        })

    records.sort(key=lambda r: r['dtstart'])
    return records


# ---------------------------------------------------------------------------
# Pivot table (pdc réalisé)
# ---------------------------------------------------------------------------
def compute_pdc_rea(df_input, out_mode='HETP'):
    df_temp = df_input.copy()
    
    pt = pd.pivot_table(
        data=df_temp, 
        index=['Cours'],  
        columns=['Activité'], 
        aggfunc=['sum'], 
        fill_value=0, 
        values=[out_mode], 
        margins=False, 
        margins_name=f'Total ({out_mode})'
    )
    # Nettoyage des multi-index colonnes
    pt.columns = [x[2] for x in pt.columns]
    pt.index.name = 'ECUE'
    return pt

# ---------------------------------------------------------------------------
# Excel Generation
# ---------------------------------------------------------------------------

COLOR_HEADER      = 'FF2E4057'  # Dark slate blue
COLOR_HEADER_FONT = 'FFFFFFFF'
COLOR_TOTAL       = 'FFFFE082'  # Amber

MODALITY_COLORS = {
    'CM':              'FFD6EAF8',  # Light blue
    'CM_Trou_ADE':     'FFFFF3CD',  # Amber (détection incertaine)
    'CM-plus':         'FFAED6F1',  # Medium blue (amphithéâtre E1/E2)
    'CM-plus_Trou_ADE':'FFFFF3CD',
    'TD':              'FFD5F5E3',  # Light green
    'TD_Trou_ADE':     'FFFFF3CD',
    'TDR':             'FFF2D7D5',  # Light red
    'TDR_Trou_ADE':    'FFFFF3CD',
    'TP':              'FFFEF9E7',  # Light yellow
    'TP_Trou_ADE':     'FFFFF3CD',
    'TP Seul':         'FFFDEBD0',  # Light orange
    'TP Seul_Trou_ADE':'FFFFF3CD',
    'CM/TD':           'FFE8DAEF',  # Light purple
    'CM/TD_Trou_ADE':  'FFFFF3CD',
    'Oraux':           'FFD7DBDD',  # Light grey
    'Oraux_Trou_ADE':  'FFFFF3CD',
    'Soutenance_Trou_ADE': 'FFFFF3CD',
    'Autre':           'FFFDFEFE',  # Near-white
}

MODALITY_ORDER = [
    'CM', 'CM_Trou_ADE', 'CM-plus', 'CM-plus_Trou_ADE',
    'CM/TD', 'CM/TD_Trou_ADE',
    'TD', 'TD_Trou_ADE', 'TDR', 'TDR_Trou_ADE',
    'TP', 'TP_Trou_ADE', 'TP Seul', 'TP Seul_Trou_ADE',
    'Oraux', 'Oraux_Trou_ADE', 'Soutenance', 'Soutenance_Trou_ADE', 'Autre',
]

# Coefficients HETD : 1h de cours × coeff = HETD
# TD=1 (référence), CM=4/3≈1.333 (2 HETP/1.5), CM-plus=5/3≈1.667 (2.5 HETP/1.5)
# TP=1/1.5≈0.667, TP Seul=1, TDR=1.5, Oraux=1/1.5, Soutenance=1/1.5
# CM/TD et Autre = non comptabilisé (0)
HETD_COEFFICIENTS_ESIEE = {
    'CM':         4 / 3,
    'CM/TD':      1.25,
    'CM-plus':    5 / 3,
    'TD':         1.0,
    'TDR':        4/3,
    'TP':         1 / 1.5,
    'TP Seul':    1.0,
    'Oraux':      1 / 1.5,
    'Soutenance': 1 / 1.5,
    'Autre':      0.0,
}

HETD_COEFFICIENTS_Univ = {
    'CM':         1.5,
    'CM/TD':      1.25,
    'CM-plus':    1.5,
    'TD':         1.0,
    'TDR':        1.5,
    'TP':         1 / 1.5,
    'TP Seul':    1.0,
    'Oraux':      1 / 1.5,
    'Soutenance': 1 / 1.5,
    'Autre':      0.0,
}

HETP_COEFFICIENTS = {
    'CM':         2,
    'CM/TD':      0,
    'CM-plus':    2.5,
    'TD':         1.5,
    'TDR':        2,
    'TP':         1,
    'TP Seul':    1.5,
    'Oraux':      1,
    'Soutenance': 1,
    'Autre':      0.0,
}


def hetd(duration_h, modality):
    """Convert hours to HETD (Heures Équivalent TD)."""
    base = modality.replace('_Trou_ADE', '')
    return duration_h * HETD_COEFFICIENTS_ESIEE.get(base, 0.0)

def hetp(duration_h, modality):
    """Convert hours to HETP (Heures Équivalent TP)."""
    base = modality.replace('_Trou_ADE', '')
    return duration_h * HETP_COEFFICIENTS.get(base, 0.0)


_thin  = Side(style='thin')
_med   = Side(style='medium')
BORDER_THIN   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_MEDIUM = Border(left=_med,  right=_med,  top=_med,  bottom=_med)


def _header_cell(cell, text):
    cell.value = text
    cell.font      = Font(bold=True, color=COLOR_HEADER_FONT)
    cell.fill      = PatternFill('solid', fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = BORDER_THIN


# -- Sheet 1: Detail --

def write_detail_sheet(ws, records):
    headers = ['Nom', 'Filiere', 'Promo', 'Date', 'Annee', 'Mois', 'Debut', 'Fin', 'Duree (h)', 'HETD (h)', 'HETP (h)', 'Lieu', 'Modalite', 'Description']
    for col, h in enumerate(headers, 1):
        _header_cell(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 30

    for i, rec in enumerate(records, 2):
        bg = MODALITY_COLORS.get(rec['modality'], MODALITY_COLORS['Autre'])
        values = [
            rec['nom'],
            rec['filiere'],
            rec['promo'],
            rec['dtstart'].strftime('%d/%m/%Y'),
            rec['dtstart'].year,
            rec['dtstart'].month,
            rec['dtstart'].strftime('%H:%M'),
            rec['dtend'].strftime('%H:%M'),
            round(rec['duration_h'], 2),
            round(hetd(rec['duration_h'], rec['modality']), 2),
            round(hetp(rec['duration_h'], rec['modality']), 2),
            rec['location'],
            rec['modality'],
            rec['description'],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = PatternFill('solid', fgColor=bg)
            cell.border = BORDER_THIN
            if col in (4, 5, 6, 7, 8, 9, 10, 11):
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif col == 14:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='top')

    col_widths = [42, 22, 20, 12, 8, 8, 8, 8, 10, 10, 10, 20, 12, 60]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    ws.freeze_panes = 'A2'


# -- Sheet 2: Recapitulatif --

def write_recap_sheet(ws, records):
    counts     = defaultdict(int)
    hours      = defaultdict(float)
    hetd_hours = defaultdict(float)
    hetp_hours = defaultdict(float)
    for rec in records:
        m = rec['modality']
        counts[m]     += 1
        hours[m]      += rec['duration_h']
        hetd_hours[m] += hetd(rec['duration_h'], m)
        hetp_hours[m] += hetp(rec['duration_h'], m)

    headers = ['Modalite', 'Coeff HETD', 'Coeff HETP', 'Nombre de seances', 'Heures totales', 'HETD totales', 'HETP totales']
    for col, h in enumerate(headers, 1):
        _header_cell(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 25

    row = 2
    total_count = 0
    total_hours = 0.0
    total_hetd  = 0.0
    total_hetp  = 0.0

    # Use fixed order, skip missing modalities
    present = [m for m in MODALITY_ORDER if m in counts]
    for m in counts:
        if m not in present:
            present.append(m)

    for m in present:
        bg    = MODALITY_COLORS.get(m, MODALITY_COLORS['Autre'])
        coeff = HETD_COEFFICIENTS_ESIEE.get(m, 0.0)
        coefftp = HETP_COEFFICIENTS.get(m, 0.0)
        vals  = [m, coeff,coefftp, counts[m], round(hours[m], 2), round(hetd_hours[m], 2), round(hetp_hours[m], 2)]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.border    = BORDER_THIN
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')
            if col == 1:
                cell.font = Font(bold=True)
        total_count += counts[m]
        total_hours += hours[m]
        total_hetd  += hetd_hours[m]
        total_hetp  += hetp_hours[m]
        row += 1

    # TOTAL row
    for col, val in enumerate(['TOTAL', '','', total_count, round(total_hours, 2), round(total_hetd, 2), round(total_hetp, 2)], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill('solid', fgColor=COLOR_TOTAL)
        cell.border    = BORDER_MEDIUM
        cell.alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')

    for col, w in enumerate([18, 14, 22, 18, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A2'


# -- Sheet 3: Avertissements --

def write_warnings_sheet(ws):
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 82

    title = ws.cell(row=1, column=1, value='Avertissements sur la detection des modalites')
    title.font      = Font(bold=True, size=13, color=COLOR_HEADER_FONT)
    title.fill      = PatternFill('solid', fgColor=COLOR_HEADER)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 28

    sub_fill = PatternFill('solid', fgColor='FF3D5A80')
    for col, text in enumerate(['Cas', 'Explication'], 1):
        cell = ws.cell(row=2, column=col, value=text)
        cell.font      = Font(bold=True, color='FFFFFFFF')
        cell.fill      = sub_fill
        cell.border    = BORDER_THIN
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    warnings = [
        ('Principe',
         'La modalite est extraite de la DERNIERE ligne non-vide du champ DESCRIPTION. '
         'Si aucune modalite n\'est trouvee dans la description, le NOM du cours (champ SUMMARY) '
         'est utilise en second recours. '
         'La priorite a la description evite les faux-positifs lies aux noms de cours '
         '(ex : "TP Capteurs") ou de salles (ex : "PER-5TP").'),
        ('Ordre de priorite',
         'CM/TD > Oraux > TDR > TP Seul > CM > TD > TP > Autre. '
         'Les patterns utilisent des limites de mots (\\b) pour eviter les faux-positifs partiels '
         '(ex : "TDSTART" ne correspond pas a "TD").'),
        ('Classe "Autre"',
         'Si la derniere ligne ne correspond a aucune modalite connue (Jury, Reunion, ...), '
         "l'evenement est classe \"Autre\". Ces heures ne sont generalement pas des heures d'enseignement."),
        ('Fuseau horaire',
         'Les heures du fichier .ics sont en UTC (suffixe Z). Ce script applique +1h (CET, heure francaise d\'hiver). '
         'En periode d\'heure d\'ete (CEST, +2h, d\'avril a octobre), les horaires affiches seraient decales '
         "d'une heure. Verifier les creneaux en periode estivale si les heures semblent incorrectes."),
        ('Encodage',
         'Le fichier .ics exporte par ADE peut contenir des caracteres mal encodes (e affiche comme Ae, etc.). '
         'Le script lit en UTF-8 avec remplacement des erreurs. '
         "En cas d'affichage incorrect, verifier l'encodage du fichier source."),
        ('Durees nulles',
         'Les evenements dont la duree calculee est nulle ou negative sont ignores silencieusement. '
         'Cela peut indiquer un probleme dans le fichier .ics source.'),
        ('Lignes vides',
         'Les lignes vides en debut et fin de description sont ignorees. '
         'La modalite est toujours extraite de la derniere ligne non-vide.'),
        ('Nouvelles modalites',
         'Si ADE introduit une nouvelle modalite non listee ici, elle sera classee "Autre". '
         'Mettre a jour MODALITY_PATTERNS dans le script si necessaire.'),
    ]

    for i, (case, explanation) in enumerate(warnings, 3):
        bg = 'FFF0F4F8' if i % 2 == 0 else 'FFFFFFFF'
        c1 = ws.cell(row=i, column=1, value=case)
        c1.font      = Font(bold=True)
        c1.fill      = PatternFill('solid', fgColor=bg)
        c1.border    = BORDER_THIN
        c1.alignment = Alignment(vertical='top', wrap_text=True)

        c2 = ws.cell(row=i, column=2, value=explanation)
        c2.fill      = PatternFill('solid', fgColor=bg)
        c2.border    = BORDER_THIN
        c2.alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[i].height = 52


def generate_excel(records, output_path):
    wb = openpyxl.Workbook()

    ws_detail = wb.active
    ws_detail.title = 'Detail'
    write_detail_sheet(ws_detail, records)

    ws_recap = wb.create_sheet('Recapitulatif')
    write_recap_sheet(ws_recap, records)

    ws_warn = wb.create_sheet('Avertissements')
    write_warnings_sheet(ws_warn)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Terminal Summary
# ---------------------------------------------------------------------------

def print_summary(records):
    counts = defaultdict(int)
    hours  = defaultdict(float)
    hetd_h = defaultdict(float)
    hetp_h = defaultdict(float)
    for rec in records:
        m = rec['modality']
        counts[m] += 1
        hours[m]  += rec['duration_h']
        hetd_h[m] += hetd(rec['duration_h'], m)
        hetp_h[m] += hetp(rec['duration_h'], m)

    present = [m for m in MODALITY_ORDER if m in counts]
    for m in counts:
        if m not in present:
            present.append(m)

    w1, w2, w3, w4 = 12, 20, 16, 14
    sep = f"+{'-'*w1}+{'-'*w2}+{'-'*w3}+{'-'*w4}+"
    print()
    print(sep)
    print(f"| {'Modalite':<{w1-2}} | {'Nombre de seances':^{w2-2}} | {'Heures totales':^{w3-2}} | {'HETD':^{w4-2}} | {'HETP':^{w4-2}} |")
    print(sep)
    for m in present:
        print(f"| {m:<{w1-2}} | {counts[m]:^{w2-2}} | {hours[m]:^{w3-3}.2f} | {hetd_h[m]:^{w4-3}.2f} | {hetp_h[m]:^{w4-3}.2f} |")
    print(sep)
    print(f"| {'TOTAL':<{w1-2}} | {sum(counts.values()):^{w2-2}} | {sum(hours.values()):^{w3-3}.2f} | {sum(hetd_h.values()):^{w4-3}.2f} | {sum(hetp_h.values()):^{w4-3}.2f} |")
    print(sep)
    print(f"  {len(records)} evenements au total\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python ade_heures.py input.ics [output.xlsx]")
        sys.exit(1)

    input_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base = input_path.rsplit('.', 1)[0] if '.' in input_path else input_path
        output_path = base + '_heures.xlsx'

    print(f"Lecture de : {input_path}")
    raw_events = parse_ics(input_path)
    print(f"  {len(raw_events)} VEVENT trouves")

    records = process_events(raw_events)
    print(f"  {len(records)} evenements traites")

    print_summary(records)

    print(f"Generation du fichier Excel : {output_path}")
    generate_excel(records, output_path)
    print("Termine.")


if __name__ == '__main__':
    main()

